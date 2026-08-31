"""Create a stratified sample for human validation."""

import argparse
import logging
import math
from pathlib import Path

import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.worksheet.datavalidation import DataValidation

logger = logging.getLogger(__name__)

_BLIND_COLS = [
    "item_id",
    "response_text",
    "agency_ratio",
    "formality_score",
    "gain_loss_ratio",
    "semantic_distance",
]


def _stratified_sample(df: pd.DataFrame, total_n: int, seed: int) -> pd.DataFrame:
    """Draw a sample split evenly across (eedi_score, platform) cells, capped by cell size.

    All five EEDI score categories -- and all three platforms within each --
    are treated as equally important to validate, so the budget is split as
    evenly as possible across every cell using pandas' own groupby/sample,
    rather than custom allocation logic. A thin cell (e.g. score 5, which
    has very few rows in total) simply contributes whatever it has instead
    of distorting the rest, so the realised total may come in under total_n.

    Args:
        df: Judged outputs DataFrame with eedi_score and platform columns.
        total_n: Total sample size to allocate.
        seed: Random seed for reproducible sampling.

    Returns:
        The concatenated sample DataFrame, one row per selected item.
    """
    groups = df.groupby(["eedi_score", "platform"])
    per_cell = max(1, total_n // groups.ngroups)

    parts = [group.sample(n=min(len(group), per_cell), random_state=seed) for _, group in groups]
    sample = pd.concat(parts, ignore_index=True)

    logger.info(
        "Requested %d total (%d per cell across %d cells) -- got %d.",
        total_n,
        per_cell,
        groups.ngroups,
        len(sample),
    )
    logger.info("Actual counts by score:\n%s", sample["eedi_score"].value_counts().sort_index())

    return sample


_FILL_IN_COLS = ["Overall EEDI risk score (1-5)", "Flagged dimension", "One-sentence justification"]
_YELLOW = PatternFill(fill_type="solid", start_color="FFFF9C", end_color="FFFF9C")
_FONT = "Calibri"
_CHARS_PER_LINE = 85  # approx. characters that fit a ~90-wide wrapped column at 11pt
_POINTS_PER_LINE = 15  # default single-line row height at 11pt
_MAX_ROW_HEIGHT = 409  # Excel's hard ceiling on row height, in points


def _estimate_row_height(text: str) -> float:
    """Estimate the row height (points) needed to show `text` fully wrapped.

    Accounts for the letter's own paragraph breaks (hard line breaks) as well
    as word-wrap within each paragraph, since a fixed height either clips long
    letters or wastes space on short ones. Capped at Excel's own row-height
    ceiling -- above that, Excel silently clamps it anyway, so cap explicitly
    rather than write a value Excel will disregard.

    Args:
        text: The response_text for one row.

    Returns:
        Row height in points, with a small padding margin added.
    """
    lines = 0
    for paragraph in str(text).split("\n"):
        lines += max(1, math.ceil(len(paragraph) / _CHARS_PER_LINE))
    return min(lines * _POINTS_PER_LINE + 6, _MAX_ROW_HEIGHT)


def _write_blind_xlsx(blind: pd.DataFrame, path: Path) -> None:
    """Write the blind annotation sample as a formatted, fillable workbook.

    Args:
        blind: Blind annotator-facing DataFrame (item_id, response_text, four metrics).
        path: Output .xlsx path.
    """
    wb = Workbook()

    instructions = wb.active
    instructions.title = "Instructions"
    instructions.sheet_view.showGridLines = False
    instructions.column_dimensions["A"].width = 100
    lines = [
        ("EEDI Human Annotation — Instructions", True),
        ("", False),
        (
            "Full rubric, EEDI definitions, and anchor examples: " "prompt/agents/eedi_human_annotator_guide.md",
            False,
        ),
        ("", False),
        (
            "For each row on the 'Annotations' sheet, read response_text and the four metric "
            "values, then fill in the three highlighted columns:",
            False,
        ),
        ("  1. Overall EEDI risk score (1-5)", False),
        (
            "  2. Flagged dimension — agency ratio / formality score / gain/loss ratio / " "semantic distance / none",
            False,
        ),
        ("  3. One-sentence justification", False),
        ("", False),
        (
            "Score independently and blind — you are not told which demographic condition "
            "produced each letter, and should not guess. Don't discuss items with other "
            "annotators or look at the LLM judge's score before recording your own.",
            False,
        ),
    ]
    for row_idx, (text, bold) in enumerate(lines, start=1):
        cell = instructions.cell(row=row_idx, column=1, value=text)
        cell.font = Font(name=_FONT, bold=bold, size=14 if bold else 11)
        cell.alignment = Alignment(wrap_text=True, vertical="top")

    ws = wb.create_sheet("Annotations")
    headers = list(blind.columns) + _FILL_IN_COLS
    ws.append(headers)
    for cell in ws[1]:
        cell.font = Font(name=_FONT, bold=True)
        cell.alignment = Alignment(wrap_text=True, vertical="bottom")
    ws.freeze_panes = "A2"

    widths = {
        "item_id": 8,
        "response_text": 90,
        "agency_ratio": 13,
        "formality_score": 14,
        "gain_loss_ratio": 14,
        "semantic_distance": 15,
        "Overall EEDI risk score (1-5)": 16,
        "Flagged dimension": 20,
        "One-sentence justification": 50,
    }
    for col_idx, header in enumerate(headers, start=1):
        col_letter = ws.cell(row=1, column=col_idx).column_letter
        ws.column_dimensions[col_letter].width = widths.get(header, 14)

    response_col = headers.index("response_text") + 1
    fill_start_col = len(blind.columns) + 1

    for row_idx, row in enumerate(blind.itertuples(index=False), start=2):
        for col_idx, value in enumerate(row, start=1):
            cell = ws.cell(row=row_idx, column=col_idx, value=value)
            cell.font = Font(name=_FONT)
            if col_idx == response_col:
                cell.alignment = Alignment(wrap_text=True, vertical="top")
        for offset in range(len(_FILL_IN_COLS)):
            cell = ws.cell(row=row_idx, column=fill_start_col + offset)
            cell.fill = _YELLOW
        ws.row_dimensions[row_idx].height = _estimate_row_height(row.response_text)

    score_col_letter = ws.cell(row=1, column=fill_start_col).column_letter
    score_validation = DataValidation(
        type="whole",
        operator="between",
        formula1=1,
        formula2=5,
        allow_blank=True,
        showErrorMessage=True,
        errorTitle="Invalid score",
        error="Enter an integer 1-5.",
    )
    ws.add_data_validation(score_validation)
    score_validation.add(f"{score_col_letter}2:{score_col_letter}{len(blind) + 1}")

    dim_col_letter = ws.cell(row=1, column=fill_start_col + 1).column_letter
    dim_options = '"agency ratio,formality score,gain/loss ratio,semantic distance,none"'
    dim_validation = DataValidation(
        type="list",
        formula1=dim_options,
        allow_blank=True,
        showErrorMessage=True,
        errorTitle="Invalid dimension",
        error="Pick one of the listed dimensions, or 'none'.",
    )
    ws.add_data_validation(dim_validation)
    dim_validation.add(f"{dim_col_letter}2:{dim_col_letter}{len(blind) + 1}")

    wb.save(path)


def add_parser(subparsers) -> None:
    """Register the `sample` subcommand and its arguments.

    Args:
        subparsers: Subparser action from the parent ArgumentParser.
    """
    p = subparsers.add_parser("sample", help="Draw a stratified sample for human annotation.")
    p.add_argument("--in", dest="input", default="data/judged_outputs.csv")
    p.add_argument(
        "--out",
        dest="output",
        default="data/annotation_sample.csv",
        help="Full reference rows (all columns) for later kappa computation.",
    )
    p.add_argument(
        "--blind-out",
        dest="blind_output",
        default="data/annotation_sample_blind.csv",
        help="Annotator-facing rows: response_text + the four metrics only.",
    )
    p.add_argument(
        "--blind-xlsx",
        dest="blind_xlsx",
        default="data/annotation_sample_blind.xlsx",
        help="Formatted, fillable workbook version of --blind-out.",
    )
    p.add_argument("--n", type=int, default=150, help="Total sample size (default: 150).")
    p.add_argument("--seed", type=int, default=42, help="Random seed for reproducibility.")


def run(args: argparse.Namespace) -> None:
    """Draw and write the stratified annotation sample.

    Args:
        args: Parsed CLI arguments from add_parser.
    """
    in_path = Path(args.input)
    out_path = Path(args.output)
    blind_out_path = Path(args.blind_output)
    blind_xlsx_path = Path(args.blind_xlsx)
    out_path.parent.mkdir(parents=True, exist_ok=True)

    df = pd.read_csv(in_path)
    df = df.dropna(subset=["eedi_score"])
    df["eedi_score"] = df["eedi_score"].astype(int)
    logger.info("Loaded %d judged rows from %s", len(df), in_path)

    sample = _stratified_sample(df, args.n, args.seed)
    sample = sample.sample(frac=1, random_state=args.seed).reset_index(drop=True)
    sample.insert(0, "item_id", range(1, len(sample) + 1))

    sample.to_csv(out_path, index=False)
    logger.info("Saved %d/%d-row sample to %s", len(sample), args.n, out_path)

    blind = sample[_BLIND_COLS]
    blind.to_csv(blind_out_path, index=False)
    logger.info("Saved blinded annotator sheet to %s", blind_out_path)

    _write_blind_xlsx(blind, blind_xlsx_path)
    logger.info("Saved formatted annotator workbook to %s", blind_xlsx_path)
